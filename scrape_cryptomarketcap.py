
from bs4 import BeautifulSoup
from urllib.error import URLError
from urllib.request import urlopen
from ast import literal_eval
from openpyxl import load_workbook
import time, argparse, pprint, xlrd, xlsxwriter, datetime


class CryptoMarketCap(object):

    def __init__(self, url, output_path):
        self.cryptoMktCap = {}  # ticker: [ Name, mkt_cap, percent_change (24), price, volume(24), circulating supply (24)]
        self.url = url
        self.output_path = output_path

    def get_CryptoMarketCap(self):
        check = True
        try:
            response = urlopen(self.url)
            html = response.read()
        except URLError as e:
            print(e)
            check = False
            time.sleep(1)
        if check:
            soup = BeautifulSoup(html, "lxml")
            for index, tr_section in enumerate(soup.find_all('tr')):
                # if index > 2:
                #     break
                if index > 0:
                    # print(tr_section)
                    name = None
                    percent_chg_1h = tr_section.find('td', class_='no-wrap percent-1h negative_change text-right')
                    if percent_chg_1h is not None:
                        percent_chg_1h = percent_chg_1h.contents[0]
                    else:
                        percent_chg_1h = tr_section.find('td', class_='no-wrap percent-1h positive_change text-right')
                        if percent_chg_1h is not None:
                            percent_chg_1h = percent_chg_1h.contents[0]

                    percent_chg_24h = tr_section.find('td', class_='no-wrap percent-24h negative_change text-right')
                    if percent_chg_24h is not None:
                        percent_chg_24h = percent_chg_24h.contents[0]
                    else:
                        percent_chg_24h = tr_section.find('td', class_='no-wrap percent-24h positive_change text-right')
                        if percent_chg_24h is not None:
                            percent_chg_24h = percent_chg_24h.contents[0]

                    percent_chg_7d = tr_section.find('td', class_='no-wrap percent-7d negative_change text-right')
                    if percent_chg_7d is not None:
                        percent_chg_7d = percent_chg_7d.contents[0]
                    else:
                        percent_chg_7d = tr_section.find('td', class_='no-wrap percent-7d positive_change text-right')
                        if percent_chg_7d is not None:
                            percent_chg_7d = percent_chg_7d.contents[0]
                        
                    currency_name = tr_section.find('a', class_='currency-name-container')
                    if currency_name is not None:
                        currency_name = currency_name.contents[0]
                    currency_symbol = tr_section.find('td', class_='text-left col-symbol')
                    if currency_symbol is not None:
                        currency_symbol = currency_symbol.contents[0]
                    mkt_cap = tr_section.find('td', class_='no-wrap market-cap text-right')
                    if mkt_cap is not None:
                        mkt_cap = str(mkt_cap.contents[0]).replace('\n', '').replace('\t', '').replace(' ', '')
                    price = tr_section.find('a', class_='price')
                    if price is not None:
                        price = price.contents[0]
                    circulating_supply_ = tr_section.find('td', class_='no-wrap text-right circulating-supply')
                    cir_supply = circulating_supply_.find('a')
                    if cir_supply is not None:
                        cir_supply = cir_supply.contents[0].replace('\n', '').replace(' ', '')
                    volume = tr_section.find('a', class_='volume')
                    if volume is not None:
                        volume = volume.contents[0]
                    
                    self.cryptoMktCap[currency_symbol] = []
                    self.cryptoMktCap[currency_symbol].append(currency_name)
                    self.cryptoMktCap[currency_symbol].append(mkt_cap)
                    self.cryptoMktCap[currency_symbol].append(price)
                    self.cryptoMktCap[currency_symbol].append(cir_supply)
                    self.cryptoMktCap[currency_symbol].append(volume)
                    self.cryptoMktCap[currency_symbol].append(percent_chg_1h)
                    self.cryptoMktCap[currency_symbol].append(percent_chg_24h)
                    self.cryptoMktCap[currency_symbol].append(percent_chg_7d)
                   
    def write_to_path(self, cryptos):
        output = self.output_path + 'cryptos.txt'
        open(output, 'w').close()
        with open(output, 'r+') as g:
            g.write(str(cryptos))

    def prep_tickers(self, path):
        with open(path, 'r') as myfile:
            data = myfile.read().replace('\n', '')

        tickers = literal_eval(data)
        return tickers

    def write_to_excel_file(self, tickers):
        print('length tickers', len(tickers))
        def create_workbook(num):
            output = self.output_path + str(num) + '_cryptos.xlsx'
            workbook = xlsxwriter.Workbook(output)
            bold = workbook.add_format({'bold': 1})
            return (workbook, bold)
        
        count = 0
        for index, key in enumerate(tickers.keys()):
            if count == 0 or count > 99:
                print('creating new workbook!', count)
                if count > 99:
                    workbook.close()
                tup = create_workbook(index)
                workbook = tup[0]
                bold = tup[1]
                count = 0
            count += 1
            # if index > 210:
            #     break
            row = 1
            col = 0
            worksheet = workbook.add_worksheet(key)
            date_format = workbook.add_format({'num_format': 'mmmm d yyyy'})
            worksheet.write('A1', 'Date', bold)
            worksheet.write('B1', 'Name', bold)
            worksheet.write('C1', 'Market Cap', bold)
            worksheet.write('D1', 'Price', bold)
            worksheet.write('E1', 'Circulating Supply (24)', bold)
            worksheet.write('F1', 'Volume (24)', bold)
            worksheet.write('G1', 'Percent Change (1h)', bold)
            worksheet.write('H1', 'Percent Change (24h)', bold)
            worksheet.write('I1', 'Percent Change (7d)', bold)
            worksheet.write_datetime(row, col, datetime.date.today(), date_format)
            col += 1
            for element in tickers[key]:
                worksheet.write(row, col, element)
                col += 1
            row += 1

        workbook.close()

    def append_data(self):
        files = ['0', '100', '200', '300', '400', '500', '600', '700', '800', '900', '1000', '1100', '1200', '1300']
        for num in files:
            file_ = self.output_path + num + '_cryptos.xlsx'
            print('************** printing to file: ', file_, ' ****************')
            wb = load_workbook(file_)
            for ws_obj in wb.worksheets:
                currency_symbol = ws_obj.title
                print('*********** printing symbol: ', currency_symbol, ' **************')
                ws = wb[currency_symbol]
                if currency_symbol in self.cryptoMktCap:
                    l = [datetime.date.today()] + self.cryptoMktCap[currency_symbol]
                    ws.append(l)
                wb.save(file_)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--OUTPUT_PATH',
        required=False,
        default='/Users/adriangarza/dev/scrape_crypto_site/',
        help='Please type a string for the path where to output (i.e. "Your_Path/write/" )')

    args = parser.parse_args()
    # url = 'https://coinmarketcap.com'
    url = 'https://coinmarketcap.com/all/views/all/'
    cryptos_dict_path = '/Users/adriangarza/dev/scrape_crypto_site/cryptos.txt'
    cmc = CryptoMarketCap(url, args.OUTPUT_PATH)
    cmc.get_CryptoMarketCap()
    cmc.write_to_path(cmc.cryptoMktCap)
    # tickers = cmc.prep_tickers(cryptos_dict_path)
    # cmc.write_to_excel_file(tickers)
    cmc.append_data()

    # set this up on your machine at home to run as a cron job at 11:59 pm everyday
    # make sure to create your excel files first and then go a head and set up as cron job
    # commenting out the correct lines to run first
    

